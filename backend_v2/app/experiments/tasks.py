"""Experiment result import.

Moved out of ``compute.tasks``, which had grown to hold this domain's tasks alongside a
dozen others. Task names and queue routing are unchanged.
"""

from __future__ import annotations

import csv
import io
import json
import uuid

from sqlalchemy import select

from ..artifacts.models import Artifact
from ..artifacts.storage import ObjectStorage
from ..core.celery_app import celery_app
from ..core.database import SessionFactory, session_scope


def _experiment_rows(filename: str, content_type: str, data: bytes) -> list[dict]:
    if content_type == "application/json" or filename.lower().endswith(".json"):
        payload = json.loads(data.decode("utf-8"))
        rows = payload.get("results") if isinstance(payload, dict) else payload
        if not isinstance(rows, list) or not all(isinstance(item, dict) for item in rows):
            raise ValueError("experiment_json_must_contain_result_objects")
        return rows
    if content_type == "text/csv" or filename.lower().endswith(".csv"):
        return list(csv.DictReader(io.StringIO(data.decode("utf-8-sig"))))
    if filename.lower().endswith(".xlsx"):
        from openpyxl import load_workbook  # type: ignore[import-untyped]

        workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        if sheet is None:
            raise ValueError("experiment_workbook_has_no_sheet")
        values = sheet.iter_rows(values_only=True)
        header_row = next(values, None)
        if header_row is None:
            raise ValueError("experiment_workbook_empty")
        headers = [str(value or "").strip() for value in header_row]
        return [dict(zip(headers, row, strict=False)) for row in values]
    raise ValueError("experiment_format_unsupported")


EXPERIMENT_IMPORT_COLUMNS = {
    "candidate_ref",
    "experiment_type",
    "pass_status",
    "value",
    "unit",
    "conclusion",
    "batch_key",
    "failure_reason",
}


def _coerce_experiment_row(row: dict, index: int) -> tuple[dict | None, dict | None]:
    """Validate one row. Returns (values, error) with exactly one of them set."""
    values = {key: row.get(key) for key in EXPERIMENT_IMPORT_COLUMNS if row.get(key) not in ("", None)}
    if not values.get("experiment_type"):
        return None, {"row": index, "column": "experiment_type", "message": "experiment_type is required"}
    raw_value = values.get("value")
    if raw_value in ("", None):
        values["value"] = None
    else:
        try:
            values["value"] = float(str(raw_value).strip())
        except ValueError:
            return None, {
                "row": index,
                "column": "value",
                "message": f"'{raw_value}' is not a number",
            }
    values["pass_status"] = values.get("pass_status") or "unknown"
    return values, None


@celery_app.task(name="bda_v2.experiment_results_import")
def experiment_results_import(artifact_id: str, dry_run: bool = False) -> dict:
    """Import assay results from a CSV/JSON/XLSX artifact.

    Rows are validated individually: a single unparseable cell no longer discards the
    whole file. ``candidate_ref`` is resolved to a real candidate so imported results
    join up with the designs they measured.
    """
    from ..candidates.models import Candidate
    from ..experiments.models import ExperimentResult

    parsed = uuid.UUID(artifact_id)
    with SessionFactory() as session:
        artifact = session.get(Artifact, parsed)
        if artifact is None or artifact.status != "available":
            return {"artifact_id": artifact_id, "status": "missing"}
        if artifact.lineage.get("experiment_import_status") == "completed":
            return {"artifact_id": artifact_id, "status": "completed", "imported": 0}
        project_id, created_by = artifact.project_id, artifact.created_by
        filename, content_type, object_key = artifact.filename, artifact.content_type, artifact.object_key
    rows = _experiment_rows(
        filename,
        content_type,
        ObjectStorage().read_bytes(object_key, max_bytes=50 * 1024 * 1024),
    )
    if len(rows) > 10000:
        raise ValueError("experiment_import_too_many_rows")

    seen_columns = {key for row in rows for key in row}
    ignored_columns = sorted(seen_columns - EXPERIMENT_IMPORT_COLUMNS)
    prepared: list[tuple[int, dict]] = []
    errors: list[dict] = []
    for index, row in enumerate(rows, start=1):
        values, error = _coerce_experiment_row(row, index)
        if error is not None:
            errors.append(error)
        else:
            prepared.append((index, values or {}))

    report = {
        "artifact_id": artifact_id,
        "imported": 0,
        "skipped": len(errors),
        "unlinked": 0,
        "errors": errors[:200],
        "ignored_columns": ignored_columns,
    }

    if dry_run:
        # Checking candidate references is the main reason to dry-run at all: an
        # unmatched ref is the defect that used to reach the database silently.
        with SessionFactory() as session:
            known = set(
                session.scalars(select(Candidate.candidate_key).where(Candidate.project_id == project_id))
            )
        unmatched = [
            {
                "row": index,
                "column": "candidate_ref",
                "message": f"no candidate '{values['candidate_ref']}' in this project",
                "severity": "warning",
            }
            for index, values in prepared
            if values.get("candidate_ref") and str(values["candidate_ref"]) not in known
        ]
        return {
            **report,
            "status": "dry_run",
            "would_import": len(prepared),
            "unlinked": len(unmatched),
            "errors": [*errors, *unmatched][:200],
        }

    created = []
    unlinked = 0
    with session_scope() as session:
        artifact = session.get(Artifact, parsed)
        if artifact is None:
            return {"artifact_id": artifact_id, "status": "missing"}
        if artifact.lineage.get("experiment_import_status") == "completed":
            return {"artifact_id": artifact_id, "status": "completed", "imported": 0}
        # candidate_key is unique per project, so a match is unambiguous.
        candidates = {
            candidate.candidate_key: candidate.id
            for candidate in session.scalars(select(Candidate).where(Candidate.project_id == project_id))
        }
        for index, values in prepared:
            reference = values.get("candidate_ref")
            candidate_id = candidates.get(str(reference)) if reference else None
            if reference and candidate_id is None:
                unlinked += 1
                errors.append(
                    {
                        "row": index,
                        "column": "candidate_ref",
                        "message": f"no candidate '{reference}' in this project; result imported unlinked",
                        "severity": "warning",
                    }
                )
            item = ExperimentResult(
                project_id=project_id,
                created_by=created_by,
                candidate_id=candidate_id,
                source_artifact_id=parsed,
                **values,
            )
            session.add(item)
            created.append(item)
        session.flush()
        artifact.lineage = {
            **artifact.lineage,
            "experiment_import_status": "completed",
            "experiment_result_ids": [str(item.id) for item in created],
            "experiment_import_report": {
                "imported": len(created),
                "skipped": report["skipped"],
                "unlinked": unlinked,
                "ignored_columns": ignored_columns,
                "errors": errors[:200],
            },
        }
        artifact.version += 1
    return {
        **report,
        "status": "completed",
        "imported": len(created),
        "unlinked": unlinked,
        "errors": errors[:200],
    }
