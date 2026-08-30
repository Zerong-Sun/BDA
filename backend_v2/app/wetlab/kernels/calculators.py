"""
蛋白质计算核心：MW、消光系数（Biopython ProtParam）、浓度、BLI 稀释规划
"""
import io
import math
import re
from dataclasses import dataclass
from typing import Any

import numpy as np
import openpyxl  # type: ignore[import-untyped]
from Bio.SeqUtils.ProtParam import ProteinAnalysis


def sanitize_seq(sequence: str) -> str:
    """清洗序列：去换行/空格/终止符/非氨基酸字符"""
    seq = sequence.upper()
    seq = re.sub(r'\s+', '', seq)
    seq = seq.replace('*', '')
    seq = re.sub(r'[^ACDEFGHIKLMNPQRSTVWY]', '', seq)
    return seq


def calc_mw(sequence: str) -> float:
    """蛋白质分子量 (Da) — 使用 Biopython ProtParam"""
    return ProteinAnalysis(sanitize_seq(sequence)).molecular_weight()


def calc_ext_coeff(sequence: str) -> dict:
    """计算还原态和氧化态消光系数 — 使用 Biopython ProtParam (Pace et al. 1995)"""
    seq = sanitize_seq(sequence)
    pa = ProteinAnalysis(seq)
    ext_red, ext_ox = pa.molar_extinction_coefficient()  # → (reduced, oxidized)
    mw = pa.molecular_weight()
    nW = seq.count("W")
    nY = seq.count("Y")
    nC = seq.count("C")
    abs_0_1pct = ext_ox / mw if mw > 0 else 0
    return {
        "mw": round(mw, 1),
        "nW": nW, "nY": nY, "nC": nC,
        "ext_red": round(ext_red, 0),
        "ext_ox": round(ext_ox, 0),
        "abs_0_1pct": round(abs_0_1pct, 4),
    }


def calc_conc(a280: float, ext_coeff: float, mw: float,
              path_length: float = 1.0) -> dict:
    """Beer-Lambert: A = ε·c·l → c"""
    if ext_coeff <= 0:
        raise ValueError("消光系数为 0，无法用 A280 定量（序列不含 W/Y/C）")
    if path_length <= 0:
        raise ValueError(f"光程必须 > 0，当前值: {path_length} cm")
    if a280 < 0:
        raise ValueError(f"A280 不能为负数，当前值: {a280}")
    molar_M = a280 / (ext_coeff * path_length)
    molar_conc_uM = molar_M * 1e6
    mass_conc_ng_uL = molar_conc_uM * mw / 1000  # 1 µM × MW/1000 = ng/µL
    return {
        "a280": a280,
        "path_length_cm": path_length,
        "epsilon": ext_coeff,
        "mw": mw,
        "molar_conc_uM": round(molar_conc_uM, 2),
        "molar_conc_nM": round(molar_conc_uM * 1e3, 2),
        "molar_conc_M": round(molar_conc_uM / 1e6, 10),
        "mass_conc_mg_mL": round(mass_conc_ng_uL / 1000, 4),
        "mass_conc_ug_mL": round(mass_conc_ng_uL, 2),
        "mass_conc_ng_uL": round(mass_conc_ng_uL, 2),
    }


# ═══════════════════════════════════════════════════════════
#  浓度单位换算 kernel（隐藏能力：6 单位互转）
# ═══════════════════════════════════════════════════════════
# canonical 基准：molar→µM，mass→ng/µL。跨 kind（摩尔↔质量）必须提供 mw (Da)：
#   molar→mass: base × mw / 1000；mass→molar: base × 1000 / mw
# 前端 static/app.js 有逐行镜像 convertConc()，改动时两边同步。

CONC_UNITS: dict[str, dict[str, Any]] = {
    "M":     {"kind": "molar", "factor": 1e6},    # → µM
    "uM":    {"kind": "molar", "factor": 1},
    "nM":    {"kind": "molar", "factor": 1e-3},
    "mg/mL": {"kind": "mass",  "factor": 1000},   # → ng/µL
    "ug/mL": {"kind": "mass",  "factor": 1},
    "ng/uL": {"kind": "mass",  "factor": 1},
}


def convert_concentration(value: float, from_unit: str, to_unit: str,
                          mw: float | None = None) -> float:
    """6 种浓度单位互转。同 kind 直接比例换算；跨 kind（摩尔↔质量）需 mw (Da)。"""
    if from_unit not in CONC_UNITS:
        raise ValueError(f"未知单位: {from_unit}")
    if to_unit not in CONC_UNITS:
        raise ValueError(f"未知单位: {to_unit}")
    f, t = CONC_UNITS[from_unit], CONC_UNITS[to_unit]
    base = value * f["factor"]  # canonical 基准（µM 或 ng/µL）
    if f["kind"] != t["kind"]:
        if not mw or mw <= 0:
            raise ValueError("跨摩尔/质量换算需要分子量 mw (Da)")
        base = base * mw / 1000 if f["kind"] == "molar" else base * 1000 / mw
    return base / t["factor"]


@dataclass
class DilutionStep:
    step: int
    conc_uM: float
    stock_vol_uL: float
    buffer_vol_uL: float
    total_vol_uL: float


def calc_dilution_series(stock_conc_uM: float, start_conc_uM: float,
                         dilution_factor: float, n_steps: int,
                         vol_per_well_uL: float,
                         extra_dead_vol_uL: float = 0.0) -> list[DilutionStep]:
    """
    BLI 梯度稀释规划 — 连续递推稀释

    每个步骤需要足够体积来：(a) 取 vol_per_well_uL 到孔中 +
    (b) 留足下一步稀释所需的母液 + (c) 死体积裕量。

    从最后一步向前递推：第 i 步总体积 = 孔体积 + 第 i+1 步总体积/稀释倍数 + 死体积。
    递推后每步总体积向上取整到 5 μL 倍数，保证移液量尽量整洁。

    Parameters
    ----------
    stock_conc_uM : 母液浓度 (μM)
    start_conc_uM : 起始最高浓度 (μM), 必须 ≤ stock_conc_uM
    dilution_factor : 稀释倍数 (如 2 表示 2 倍稀释)
    n_steps : 梯度步数
    vol_per_well_uL : 每孔所需体积 (μL)
    extra_dead_vol_uL : 额外死体积 (μL), 用于保证移液准确

    Returns
    -------
    list of DilutionStep
    """
    if stock_conc_uM <= 0:
        raise ValueError(f"母液浓度必须 > 0，当前值: {stock_conc_uM} μM")
    if start_conc_uM <= 0:
        raise ValueError(f"起始浓度必须 > 0，当前值: {start_conc_uM} μM")
    if start_conc_uM > stock_conc_uM:
        raise ValueError(
            f"起始浓度 ({start_conc_uM} μM) 不能超过母液浓度 ({stock_conc_uM} μM)")
    if dilution_factor <= 1:
        raise ValueError(f"稀释倍数必须 > 1，当前值: {dilution_factor}")
    if n_steps < 1:
        raise ValueError(f"梯度步数必须 ≥ 1，当前值: {n_steps}")
    if vol_per_well_uL <= 0:
        raise ValueError(f"每孔体积必须 > 0，当前值: {vol_per_well_uL} μL")

    # 从后向前递推，计算第 0 步所需最大体积
    total_last = vol_per_well_uL + extra_dead_vol_uL
    total_cur = total_last
    for _ in range(n_steps - 1):
        total_cur = vol_per_well_uL + total_cur / dilution_factor + extra_dead_vol_uL
    # 向上取整到 100 μL，所有步骤统一体积
    uniform_total = math.ceil(total_cur / 100) * 100

    steps = []
    for i in range(n_steps):
        target_conc = start_conc_uM / (dilution_factor ** i)
        total_needed = uniform_total

        if i == 0:
            # 第一步: 直接从母液配制
            stock_vol = total_needed * target_conc / stock_conc_uM
            buffer_vol = total_needed - stock_vol
        else:
            # 后续步骤: 从上一步的剩余液中取 total_needed / factor
            stock_vol = total_needed / dilution_factor
            buffer_vol = total_needed - stock_vol

        steps.append(DilutionStep(
            step=i + 1,
            conc_uM=round(target_conc, 4),
            stock_vol_uL=round(stock_vol, 2),
            buffer_vol_uL=round(buffer_vol, 2),
            total_vol_uL=round(total_needed, 2),
        ))
    return steps


# ═══════════════════════════════════════════════════════════
#  酶活动力学计算
# ═══════════════════════════════════════════════════════════

ROW_ORDER = "ABCDEFGH"


def parse_tecan_xlsx(source: str | bytes) -> dict:
    """解析 TECAN Spark xlsx，返回 {meta: {...}, wells: {A1: {times:[], od:[]}, ...}}

    接受路径或原始字节。BDA 的上传落在对象存储里，服务层取到字节后直接喂进来，
    不会先写临时文件。
    """
    handle: object = io.BytesIO(source) if isinstance(source, bytes) else source
    wb = openpyxl.load_workbook(handle, data_only=True)
    ws = wb.active
    meta: dict[str, Any] = {"sample": "", "wavelength": "", "temps": []}
    wells: dict[str, dict[str, Any]] = {}

    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v is None:
            continue
        label = str(v).strip().rstrip(":")

        if label == "Name":
            meta["sample"] = str(ws.cell(r, 2).value or "")
        elif label == "Measurement wavelength":
            meta["wavelength"] = ws.cell(r, 5).value
        elif label == "Target temperature":
            meta["target_temp"] = ws.cell(r, 5).value

        elif label == "Cycle Nr.":
            time_s = float(ws.cell(r - 2, 2).value)
            meta.setdefault("temps", []).append(time_s)
            header_row = r + 1
            col_map = {}
            for c in range(2, 14):
                vh = ws.cell(header_row, c).value
                if vh is not None:
                    col_map[c] = int(vh)
            for rr in range(header_row + 1, header_row + 9):
                rl = str(ws.cell(rr, 1).value or "").strip()
                if rl not in ROW_ORDER:
                    continue
                for c, wcol in col_map.items():
                    vv = ws.cell(rr, c).value
                    if vv is None or str(vv).strip() == "":
                        continue
                    key = f"{rl}{wcol}"
                    wells.setdefault(key, {"times": [], "od": []})
                    wells[key]["times"].append(time_s)
                    wells[key]["od"].append(float(vv))
    wb.close()
    return {"meta": meta, "wells": {k: v for k, v in sorted(wells.items())}}


def fit_kinetics(times: list, od: list) -> dict:
    """线性拟合 → ΔOD/min、R²"""
    t = np.asarray(times, float)
    od_arr = np.asarray(od, float)
    n = len(t)
    if n < 2:
        return {"slope": None, "intercept": None, "r2": None, "n": n}
    k, b = np.polyfit(t, od_arr, 1)
    dod_min = round(float(k * 60), 6)
    if n > 2:
        pred = k * t + b
        ss_res = float(np.sum((od_arr - pred) ** 2))
        ss_tot = float(np.sum((od_arr - od_arr.mean()) ** 2))
        r2 = round(float(1 - ss_res / ss_tot), 4) if ss_tot > 0 else None
    else:
        r2 = None
    return {"slope": dod_min, "intercept": round(float(b), 6), "r2": r2, "n": n}


def _bg_priority(refs_present):
    """背景孔优先级：阴性(neg) > 空白(blank)。两者并存只扣阴性——
    空白是缓冲液基线，混进背景均值再扣回自己身上 = 被多扣一次。无背景返回 None。"""
    if "neg" in refs_present:
        return ("neg",)
    if "blank" in refs_present:
        return ("blank",)
    return None


def sub_blank(wells: dict, enabled: bool = True):
    """扣除背景信号（sub_blank）：逐时间点减背景孔均值 OD，背景自身归零。
    背景 = 阴性(neg) 孔（有样本无酶，捕获真实反应背景）；无 neg 回退空白(blank) 孔（缓冲液基线）。
    时间值匹配（而非索引），兼容个别孔缺测点错位。
    返回 (new_wells, mean_neg)：new_wells 每个孔 od 被替换为扣减后的新列表；
    mean_neg 为 {time: mean_od}，未启用或无背景孔时返回 None。
    注意：new_wells 是浅拷贝——受影响孔的 dict 为新对象但 fit 等嵌套结构与原对象共享引用；
    未受影响孔直接复用原 dict，调用方不要深改嵌套结构。"""
    if not enabled:
        return wells, None
    bg_refs = _bg_priority({wd.get("ref") for wd in wells.values()})
    if not bg_refs:
        return wells, None
    buckets: dict[Any, list[Any]] = {}
    for wd in wells.values():
        if wd.get("ref") in bg_refs and wd.get("times") and wd.get("od"):
            for t, o in zip(wd["times"], wd["od"], strict=True):
                buckets.setdefault(t, []).append(o)
    if not buckets:
        return wells, None
    mean_neg = {t: sum(v) / len(v) for t, v in buckets.items()}
    new_wells = {}
    for wid, wd in wells.items():
        if wd.get("times") and wd.get("od"):
            nd = dict(wd)
            nd["od"] = [o - mean_neg.get(t, 0.0) for t, o in zip(wd["times"], wd["od"], strict=True)]
            new_wells[wid] = nd
        else:
            new_wells[wid] = wd
    return new_wells, mean_neg


def align_wells(wells: dict, align_start: bool = False, align_end: bool = False):
    """对齐起始/终止值：均值只统计样品/阳性孔（阴性/空白是参考，sub_blank 后≈0）。
    位移只作用于样品/阳性孔。返回 (new_wells, avg_first, avg_last)。
    注意：与 sub_blank 同为浅拷贝语义——受影响孔为新 dict、嵌套结构与原对象共享引用。"""
    all_first_od = []
    all_last_od = []
    if align_start or align_end:
        for wd in wells.values():
            if wd.get("ref") in ("blank", "neg"):
                continue
            if wd.get("od") and len(wd["od"]) > 0:
                all_first_od.append(wd["od"][0])
                all_last_od.append(wd["od"][-1])
    avg_first = sum(all_first_od) / len(all_first_od) if all_first_od else 0.0
    avg_last = sum(all_last_od) / len(all_last_od) if all_last_od else 0.0
    new_wells = {}
    for wid, wd in wells.items():
        ref = wd.get("ref", "")
        od_vals = list(wd.get("od") or [])
        if od_vals and ref not in ("blank", "neg") and (align_start or align_end):
            if align_start:
                od_vals = [v + (avg_first - od_vals[0]) for v in od_vals]
            if align_end:
                od_vals = [v + (avg_last - od_vals[-1]) for v in od_vals]
            nd = dict(wd)
            nd["od"] = od_vals
            new_wells[wid] = nd
        else:
            new_wells[wid] = wd
    return new_wells, avg_first, avg_last


def aggregate_groups(wells_data: dict):
    """把带非空 group 的孔按组聚合为平均曲线（逐时间点 OD 均值 ± SD）。

    供动力学曲线「同组孔取平均 + 误差棒」使用。逐时间点按时间值匹配聚合
    （round 到 6 位，兼容个别孔缺测点错位）；时间单位保持输入（秒），
    调用方作图时再转分钟。

    返回 (groups, singles)：
      groups: list[dict]，每项 = {label, times, od, err, n, conc_ng_ml, mean_slope}
              times=时间点列表（升序），od=逐点均值，
              err=逐点样本标准差（ddof=1，该点仅 1 孔时为 0），
              n=组内成员孔数，conc_ng_ml=组内一致的浓度（不一致则 None），
              mean_slope=成员 fit 斜率（slope_corrected 优先，回退 slope）的非空均值；
      singles: list[well_id]，无 group 的孔 + 组内仅 1 孔的孔（退化为单孔绘制）。"""
    groups: dict[Any, list[Any]] = {}
    singles = []
    for wid, wd in wells_data.items():
        if not isinstance(wd, dict) or not wd.get("times") or not wd.get("od"):
            continue
        g = (wd.get("group") or "").strip()
        if not g:
            singles.append(wid)
        else:
            groups.setdefault(g, []).append(wid)

    out = []
    for g, member_ids in sorted(groups.items()):
        if len(member_ids) < 2:
            singles.extend(member_ids)  # 单孔组退化为普通孔
            continue
        buckets: dict[Any, list[Any]] = {}  # time -> [od...]
        fit_slopes = []
        concs = set()
        for wid in member_ids:
            wd = wells_data[wid]
            for t, od in zip(wd["times"], wd["od"], strict=True):
                if od is None:
                    continue
                key = round(float(t), 6)
                buckets.setdefault(key, []).append(float(od))
            fit = wd.get("fit") or {}
            sl = fit.get("slope_corrected")
            if sl is None:
                sl = fit.get("slope")
            if sl is not None:
                fit_slopes.append(float(sl))
            if wd.get("conc_ng_ml"):
                concs.add(wd["conc_ng_ml"])
        times = sorted(buckets)
        od = []
        err = []
        for t in times:
            vals = buckets[t]
            od.append(float(np.mean(vals)))
            err.append(float(np.std(vals, ddof=1)) if len(vals) >= 2 else 0.0)
        out.append({
            "label": g,
            "times": times,
            "od": od,
            "err": err,
            "n": len(member_ids),
            "conc_ng_ml": list(concs)[0] if len(concs) == 1 else None,
            "mean_slope": float(np.mean(fit_slopes)) if fit_slopes else None,
        })
    return out, singles


def snap_ylim(values, pad: float = 0.06):
    """纵轴范围：数据范围外扩 pad 后按数量级取整到整刻度，避免难看的自动刻度。
    返回 (lo, hi)；values 为空返回 None。"""
    if not values:
        return None
    lo, hi = min(values), max(values)
    span = max(hi - lo, 1e-9)
    lo -= pad * span
    hi += pad * span
    step = 10 ** (int(np.floor(np.log10(span))) - 1)
    return float(np.floor(lo / step) * step), float(np.ceil(hi / step) * step)


def correct_slopes(fits: dict, refs: dict) -> tuple[dict, dict | None]:
    """速率级背景扣除：以背景孔斜率均值校正样品/阳性/阴性自身。
    背景 = 阴性(neg) 优先，无 neg 回退空白(blank)（两者并存只扣阴性）。
    空白(blank) 孔是缓冲液基线——信号层已被 sub_blank 扣平，速率层再扣 = 被多扣一次成负值，故不做速率校正。
    fits: {well_id: fit_kinetics 输出}；refs: {well_id: ref}。
    返回 (out, bg)：out 为非空白孔补 blank_corrected / slope_corrected 的新 fits；
    bg = {"avg": 背景均值, "count": 背景孔数}，无背景返回 None。"""
    bg_refs = _bg_priority(set(refs.values()))
    bg_slopes = []
    if bg_refs:
        for wid, fit in fits.items():
            if refs.get(wid) in bg_refs and fit.get("slope") is not None:
                bg_slopes.append(fit["slope"])
    bg = {"avg": sum(bg_slopes) / len(bg_slopes), "count": len(bg_slopes)} if bg_slopes else None
    out = {}
    for wid, fit in fits.items():
        nf = dict(fit)
        if fit.get("slope") is not None and refs.get(wid) != "blank":
            nf["blank_corrected"] = bool(bg)
            if bg:
                nf["slope_corrected"] = round(fit["slope"] - bg["avg"], 6)
        out[wid] = nf
    return out, bg
